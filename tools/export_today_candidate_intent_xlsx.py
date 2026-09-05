from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\codex\outputs\football_odds_trader")
LEDGER = BASE / "ledger"
DASHBOARD = BASE / "dashboard" / "index.html"
EXPORTS = BASE / "exports"


HEADER_FILL = PatternFill("solid", fgColor="1F6FAE")
HEADER_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Microsoft YaHei", color="003A5D")
RED_FONT = Font(name="Microsoft YaHei", color="D91E18", bold=True)
MUTED_FILL = PatternFill("solid", fgColor="EAF4FC")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
BORDER = Border(
    left=Side(style="thin", color="B7D0E6"),
    right=Side(style="thin", color="B7D0E6"),
    top=Side(style="thin", color="B7D0E6"),
    bottom=Side(style="thin", color="B7D0E6"),
)


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def latest_file(root: Path, pattern: str) -> Path | None:
    files = sorted(root.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def strip_html(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("&nbsp;", " ")
    return re.sub(r"\s+", " ", text).strip()


def load_cards() -> list[dict[str, Any]]:
    text = DASHBOARD.read_text(encoding="utf-8")
    match = re.search(r"const cardsData = (\[.*?\]);\n", text, re.S)
    if not match:
        raise RuntimeError(f"cardsData not found in {DASHBOARD}")
    return json.loads(match.group(1))


def parse_time(value: str) -> datetime:
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{2})", value or "")
    if not match:
        return datetime.max
    y, mo, d, h, mi = map(int, match.groups())
    return datetime(y, mo, d, h, mi)


def side_text(value: str) -> str:
    return {"home": "主队", "away": "客队", "upper": "上盘", "lower": "下盘"}.get(value or "", value or "")


def history_maps() -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]], dict[tuple[str, str], dict[str, str]]]:
    summary_path = latest_file(LEDGER, "asian_intent_history_summary_*.csv")
    tag_map: dict[str, dict[str, str]] = {}
    line_tag_map: dict[str, dict[str, str]] = {}
    if summary_path:
        for row in read_csv(summary_path):
            group_type = row.get("分组类型", "")
            group = row.get("分组", "")
            if group_type == "候选标签":
                tag_map[group] = row
            elif group_type == "盘口档位+候选标签":
                line_tag_map[group] = row

    micro_path = latest_file(LEDGER, "micro_region_tag_edge_*.csv")
    micro_map: dict[tuple[str, str], dict[str, str]] = {}
    if micro_path:
        for row in read_csv(micro_path):
            micro_map[(row.get("微观板块", ""), row.get("候选标签", ""))] = row
    return tag_map, line_tag_map, micro_map


def history_cols(card: dict[str, Any], tag_map: dict[str, dict[str, str]], line_tag_map: dict[str, dict[str, str]], micro_map: dict[tuple[str, str], dict[str, str]]) -> dict[str, str]:
    tag = strip_html(card.get("intent_tag"))
    line = strip_html(card.get("intent_line_bucket"))
    micro = strip_html(card.get("micro_region"))
    tag_row = tag_map.get(tag, {})
    line_tag_row = line_tag_map.get(f"{line} / {tag}") or line_tag_map.get(f"{line}/{tag}") or {}
    micro_row = micro_map.get((micro, tag), {})
    return {
        "标签样本": tag_row.get("样本", ""),
        "标签正向胜率": tag_row.get("有效胜率", ""),
        "标签正向盈亏": tag_row.get("均注盈亏", ""),
        "标签反向胜率": tag_row.get("反向有效胜率", ""),
        "标签反向盈亏": tag_row.get("反向均注盈亏", ""),
        "标签建议方向": tag_row.get("建议方向", ""),
        "同盘口+标签样本": line_tag_row.get("样本", ""),
        "同盘口+标签正向胜率": line_tag_row.get("有效胜率", ""),
        "同盘口+标签正向盈亏": line_tag_row.get("均注盈亏", ""),
        "同盘口+标签反向胜率": line_tag_row.get("反向有效胜率", ""),
        "同盘口+标签反向盈亏": line_tag_row.get("反向均注盈亏", ""),
        "同盘口+标签建议方向": line_tag_row.get("建议方向", ""),
        "微观样本": micro_row.get("样本数", ""),
        "微观正向胜率": micro_row.get("正向有效胜率", ""),
        "微观正向盈亏": micro_row.get("正向均注盈亏", ""),
        "微观反向胜率": micro_row.get("反向有效胜率", ""),
        "微观反向盈亏": micro_row.get("反向均注盈亏", ""),
        "微观优先方向": micro_row.get("微观优先方向", ""),
        "微观建议动作": micro_row.get("建议动作", ""),
        "风控状态": micro_row.get("风控状态", ""),
    }


def card_to_row(card: dict[str, Any], tag_map: dict[str, dict[str, str]], line_tag_map: dict[str, dict[str, str]], micro_map: dict[tuple[str, str], dict[str, str]]) -> dict[str, Any]:
    hist = history_cols(card, tag_map, line_tag_map, micro_map)
    forward_team = strip_html(card.get("intent_forward_team"))
    reverse_team = strip_html(card.get("intent_reverse_team"))
    return {
        "列表日": card.get("date", ""),
        "北京时间": strip_html(card.get("display_time") or card.get("time")),
        "赛事": strip_html(card.get("league")),
        "中文比赛": strip_html(card.get("display_match") or card.get("match")),
        "状态": strip_html(card.get("display_status") or card.get("status")),
        "比分": strip_html(card.get("display_score") or card.get("score")),
        "微观板块": strip_html(card.get("micro_region")),
        "排名/阶段": strip_html(card.get("rank")),
        "亚盘完整": "是" if card.get("ah_ok") else "否",
        "欧赔完整": "是" if card.get("euro_ok") else "否",
        "大小球完整": "是" if card.get("total_ok") else "否",
        "亚盘开盘->即时": strip_html(card.get("ah")),
        "欧赔开盘->即时": strip_html(card.get("euro")),
        "欧赔去水": strip_html(card.get("euro_devig")),
        "大小球开盘->即时": strip_html(card.get("total")),
        "候选标签": strip_html(card.get("intent_tag")),
        "盘口档位": strip_html(card.get("intent_line_bucket")),
        "上盘方": strip_html(card.get("intent_upper_team")),
        "上盘主客": side_text(str(card.get("intent_upper_side") or "")),
        "上盘水位": card.get("intent_upper_water"),
        "下盘方": strip_html(card.get("intent_lower_team")),
        "下盘水位": card.get("intent_lower_water"),
        "正向候选方": forward_team or "无方向/待筛",
        "正向候选盘向": side_text(str(card.get("intent_forward_side") or "")),
        "正向候选水位": card.get("intent_forward_water"),
        "反向候选方": reverse_team or "无方向/待筛",
        "反向候选水位": card.get("intent_reverse_water"),
        "是否已冻结可投": "是" if card.get("frozen_bettable") else "否",
        "冻结动作": strip_html(card.get("frozen_bettable_action")),
        "冻结投注球队": strip_html(card.get("frozen_bettable_team")),
        "冻结投注盘向": strip_html(card.get("frozen_bettable_side")),
        "冻结水位": card.get("frozen_bettable_water"),
        "冻结综合胜率": card.get("frozen_bettable_rate"),
        "冻结通过阈值": card.get("frozen_bettable_threshold"),
        "当前处理口径": strip_html(card.get("final_action")),
        "模拟方向字段": strip_html(card.get("pick")),
        "错误/缺口字段": strip_html(card.get("error")),
        "亚盘意图文字": strip_html(card.get("asian_intent")),
        "基本面拉力": strip_html(card.get("pull")),
        "伤停": strip_html(card.get("injury")),
        "首发": strip_html(card.get("lineup")),
        "近况来源": strip_html(card.get("form_source")),
        "H2H来源": strip_html(card.get("h2h_source")),
        "赢盘/输盘来源": strip_html(card.get("handicap_record_source")),
        "资金流": strip_html(card.get("flow")),
        "流动性": strip_html(card.get("liquidity")),
        "数据完整性": strip_html(card.get("data_complete_detail")),
        "盘口源": strip_html(card.get("odds_source")),
        "比赛ID": strip_html(card.get("match_id")),
        "模拟ID": strip_html(card.get("sim_id")),
        **hist,
    }


def write_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        tag = str(row[headers.index("候选标签")].value) if "候选标签" in headers else ""
        ah_ok = str(row[headers.index("亚盘完整")].value) if "亚盘完整" in headers else "是"
        if ah_ok == "否":
            for cell in row:
                cell.fill = BAD_FILL
        elif "平衡盘" in tag:
            for cell in row:
                cell.fill = WARN_FILL
        elif tag and "不完整" not in tag:
            for cell in row:
                cell.fill = GOOD_FILL
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=min(ws.max_row, 80)):
            for item in cell:
                max_len = max(max_len, len(str(item.value or "")) // 2)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 42)
    ws.sheet_view.showGridLines = False


def summary_rows(cards: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    tag_counter = Counter(strip_html(c.get("intent_tag")) or "空" for c in cards)
    tag_rows = [{"候选标签": k, "今日场数": v} for k, v in tag_counter.most_common()]

    combo: defaultdict[tuple[str, str], int] = defaultdict(int)
    for c in cards:
        combo[(strip_html(c.get("micro_region")) or "未识别", strip_html(c.get("intent_tag")) or "空")] += 1
    micro_rows = [{"微观板块": k[0], "候选标签": k[1], "今日场数": v} for k, v in sorted(combo.items(), key=lambda item: (-item[1], item[0][0], item[0][1]))]
    return tag_rows, micro_rows


def export(target_date: str) -> Path:
    cards = [c for c in load_cards() if c.get("date") == target_date]
    if not cards:
        raise RuntimeError(f"no cards found for date={target_date}")
    cards.sort(key=lambda c: (parse_time(str(c.get("display_time") or c.get("time") or "")), strip_html(c.get("league")), strip_html(c.get("display_match"))))
    tag_map, line_tag_map, micro_map = history_maps()
    all_rows = [card_to_row(c, tag_map, line_tag_map, micro_map) for c in cards]
    ah_rows = [r for r in all_rows if r["亚盘完整"] == "是"]
    directional_rows = [r for r in ah_rows if r["正向候选方"] != "无方向/待筛"]
    missing_rows = [r for r in all_rows if r["亚盘完整"] != "是"]
    tag_rows, micro_rows = summary_rows(cards)

    main_headers = [
        "列表日",
        "北京时间",
        "赛事",
        "中文比赛",
        "状态",
        "比分",
        "微观板块",
        "排名/阶段",
        "亚盘完整",
        "亚盘开盘->即时",
        "欧赔开盘->即时",
        "欧赔去水",
        "大小球开盘->即时",
        "候选标签",
        "盘口档位",
        "上盘方",
        "上盘主客",
        "上盘水位",
        "下盘方",
        "下盘水位",
        "正向候选方",
        "正向候选盘向",
        "正向候选水位",
        "反向候选方",
        "反向候选水位",
        "标签样本",
        "标签正向胜率",
        "标签正向盈亏",
        "标签反向胜率",
        "标签反向盈亏",
        "标签建议方向",
        "微观样本",
        "微观正向胜率",
        "微观正向盈亏",
        "微观反向胜率",
        "微观反向盈亏",
        "微观优先方向",
        "微观建议动作",
        "风控状态",
        "同盘口+标签样本",
        "同盘口+标签正向胜率",
        "同盘口+标签正向盈亏",
        "同盘口+标签反向胜率",
        "同盘口+标签反向盈亏",
        "同盘口+标签建议方向",
        "资金流",
        "流动性",
        "伤停",
        "首发",
        "近况来源",
        "H2H来源",
        "赢盘/输盘来源",
        "数据完整性",
        "当前处理口径",
        "比赛ID",
        "模拟ID",
    ]
    all_headers = main_headers + ["欧赔完整", "大小球完整", "盘口源", "模拟方向字段", "错误/缺口字段", "亚盘意图文字", "基本面拉力"]
    missing_headers = [
        "列表日",
        "北京时间",
        "赛事",
        "中文比赛",
        "状态",
        "比分",
        "微观板块",
        "欧赔开盘->即时",
        "欧赔去水",
        "候选标签",
        "数据完整性",
        "错误/缺口字段",
        "比赛ID",
        "模拟ID",
    ]

    EXPORTS.mkdir(parents=True, exist_ok=True)
    snapshot = latest_file(BASE / "raw" / "titan007" / target_date.replace("-", ""), "*titan007_odds_snapshot.csv")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = EXPORTS / f"today_candidate_intent_{target_date}_{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "导读"
    intro = [
        ["字段", "值"],
        ["列表日", target_date],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S 北京时间")],
        ["今日总比赛", len(all_rows)],
        ["亚盘完整候选", len(ah_rows)],
        ["可判方向候选", len(directional_rows)],
        ["盘口缺失", len(missing_rows)],
        ["dashboard来源", str(DASHBOARD)],
        ["今日模拟CSV", str(LEDGER / f"{target_date}_titan007_simulations.csv")],
        ["赔率快照", str(snapshot or "")],
        ["说明", "候选意图不等同真实下注；真实下注需继续通过标签历史、微观组合、水位阈值、同盘口否决和风控。"],
    ]
    for row in intro:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90
    ws.sheet_view.showGridLines = False

    write_sheet(wb.create_sheet("今日候选意图"), ah_rows, main_headers)
    write_sheet(wb.create_sheet("可判方向候选"), directional_rows, main_headers)
    write_sheet(wb.create_sheet("盘口缺失"), missing_rows, missing_headers)
    write_sheet(wb.create_sheet("全部今日比赛"), all_rows, all_headers)
    write_sheet(wb.create_sheet("标签汇总"), tag_rows, ["候选标签", "今日场数"])
    write_sheet(wb.create_sheet("微观标签汇总"), micro_rows, ["微观板块", "候选标签", "今日场数"])

    tag_history_rows = list(tag_map.values())
    write_sheet(
        wb.create_sheet("标签历史EV"),
        tag_history_rows,
        ["分组类型", "分组", "样本", "有效胜率", "均注盈亏", "均注ROI", "反向有效胜率", "反向均注盈亏", "反向均注ROI", "平均意图水位", "建议方向"],
    )
    micro_history_rows = list(micro_map.values())
    write_sheet(
        wb.create_sheet("微观标签EV"),
        micro_history_rows,
        [
            "统计日期",
            "微观板块",
            "候选标签",
            "样本数",
            "正向有效胜率",
            "正向均注盈亏",
            "反向有效胜率",
            "反向均注盈亏",
            "微观优先方向",
            "标签优先方向",
            "贝叶斯综合胜率",
            "风控状态",
            "建议动作",
        ],
    )
    line_tag_rows = list(line_tag_map.values())
    write_sheet(
        wb.create_sheet("同盘口标签EV"),
        line_tag_rows,
        ["分组类型", "分组", "样本", "有效胜率", "均注盈亏", "均注ROI", "反向有效胜率", "反向均注盈亏", "反向均注ROI", "平均意图水位", "建议方向"],
    )

    wb.save(out)
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"), help="Titan007 list date, YYYY-MM-DD")
    args = parser.parse_args()
    out = export(args.date)
    print(out)


if __name__ == "__main__":
    main()
